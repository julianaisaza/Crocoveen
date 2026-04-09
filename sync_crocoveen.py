#!/usr/bin/env python3
"""
sync_crocoveen.py
─────────────────────────────────────────────────────────────────
Lee todos los archivos *Crocoveen*.xlsx de la carpeta donde vive
este script y actualiza en crocoveen.html los campos:
  • towers[].sold   (Vendidas)
  • S.rhythm        (ritmo mensual de ventas)
  • S.deliveries    (cronograma de entregas)

Ejecución:
    python3 sync_crocoveen.py              # usa la carpeta del script
    python3 sync_crocoveen.py /ruta/otra  # ruta explícita
"""

import openpyxl, glob, re, os, sys, json
from datetime import datetime

# ── 1. Configuración ────────────────────────────────────────────
FOLDER       = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(os.path.abspath(__file__))
EXCEL_FOLDER = os.path.dirname(FOLDER)   # carpeta padre: donde viven los Excel de ventas
HTML         = os.path.join(FOLDER, "crocoveen.html")

MO = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
      'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}

def norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

# ── 2. Tabla nombre-torre → ID de app ───────────────────────────
# Clave: (keyword_del_archivo, nombre_normalizado_de_torre)
TOWER_MAP = {
    # Country Life y Verano
    ('country','tgct1a'): 'C-T1A', ('country','tgct1b'): 'C-T1B',
    ('country','tgct2a'): 'C-T2A', ('country','tgct2b'): 'C-T2B',
    ('country','veranot1a'): 'V-T1A', ('country','veranot1b'): 'V-T1B',
    ('country','veranot2a'): 'V-T2A', ('country','veranot2b'): 'V-T2B',
    # Zitizen
    ('zitizen','t1'): 'Z-T1', ('zitizen','t2'): 'Z-T2',
    # Primavera
    ('primavera','t1'): 'P-T1', ('primavera','t2'): 'P-T2', ('primavera','t3'): 'P-T3',
    # Mágica (usa keyword 'gica' para cubrir la ó)
    ('gica','t2a'): 'M-T2A', ('gica','t2b'): 'M-T2B',
    ('gica','t3a'): 'M-T3A', ('gica','t3b'): 'M-T3B',
    ('gica','t4a'): 'M-T4A', ('gica','t4b'): 'M-T4B', ('gica','t4c'): 'M-T4C',
    # Bosketo
    ('bosketo','t1'): 'B-T1', ('bosketo','t2a'): 'B-T2A',
    ('bosketo','t2b'): 'B-T2B', ('bosketo','t3'): 'B-T3',
    # Camino Verde
    ('camino','t3'): 'CV-T3', ('camino','t4'): 'CV-T4',
    # Ambarte
    ('ambarte','t1a'): 'A-T1A', ('ambarte','t1b'): 'A-T1B', ('ambarte','t1c'): 'A-T1C',
    # La Vida es Bella (nombre normalizado empieza con 'lvb2t')
    ('bella','lvb2t1'): 'L-T1', ('bella','lvb2t2'): 'L-T2',
    ('bella','lvb2t3'): 'L-T3', ('bella','lvb2t4'): 'L-T4',
    ('bella','lvb2t5'): 'L-T5', ('bella','lvb2t6'): 'L-T6',
    # Summit – la hoja se llama UAU5
    ('uau5','t1'): 'S-T1', ('uau5','t2'): 'S-T2',
    ('uau5','t1s2'): 'S-T1S2', ('uau5','t1-s2'): 'S-T1S2',
    ('summit','t1'): 'S-T1', ('summit','t2'): 'S-T2',
    # Summit Grand
    ('summitgrand','t1'): 'SG-T1', ('summitgrand','t2'): 'SG-T2', ('summitgrand','t3'): 'SG-T3',
    ('grand','t1'): 'SG-T1', ('grand','t2'): 'SG-T2', ('grand','t3'): 'SG-T3',
    # Lúmina (keyword 'lmina' = norm de 'Lúmina'; 'bosques' = nombre de hoja)
    ('lmina','t1'): 'Lu-T1', ('lmina','t2'): 'Lu-T2', ('lmina','t3'): 'Lu-T3',
    ('bosques','t1'): 'Lu-T1', ('bosques','t2'): 'Lu-T2', ('bosques','t3'): 'Lu-T3',
}

NOT_TOWER = {'total','ventas','rph','torre','locales','condiciones',
             'cronentregas','planadepagos','actualizacion'}

def resolve_tower_id(file_keys, tower_name_raw):
    tn = norm(tower_name_raw)
    for fk in file_keys:
        hit = TOWER_MAP.get((fk, tn))
        if hit:
            return hit
    # Partial key match (exact normalized name)
    for (fk, tk), tid in TOWER_MAP.items():
        if tk == tn and any(fk in k for k in file_keys):
            return tid
    # Prefix match: handles names like 'LVB 2 T1\nARMONIA' → norm = 'lvb2t1armonia'
    # where the key is 'lvb2t1'. Only match if the char after the prefix is NOT a digit
    # (to avoid 'lvb2t10' matching key 'lvb2t1')
    for (fk, tk), tid in TOWER_MAP.items():
        if tn.startswith(tk) and any(fk in k for k in file_keys):
            suffix = tn[len(tk):]
            if not suffix or not suffix[0].isdigit():
                return tid
    return None

# ── 3. Mapeo columna → mes (YYYY-MM) ───────────────────────────
def build_col_month(hdr_vals, data_start_idx):
    """hdr_vals: tuple de 1 fila; data_start_idx: índice 0-based donde empiezan los meses."""
    col_month = {}
    i = data_start_idx
    while i < len(hdr_vals):
        v = hdr_vals[i]
        col = i + 1  # 1-based
        if v is None:
            i += 1
        elif isinstance(v, datetime):
            col_month[col] = v.strftime('%Y-%m')
            i += 1
        elif isinstance(v, (int, float)) and not isinstance(v, bool):
            i += 1  # número de año, saltar
        elif isinstance(v, str):
            clean = v.strip().replace(' ', '')
            # Formato 'Abr-Jun-26', 'Ene-Mar27', etc.
            m = re.match(r'([A-Za-z]{3})[-]?[A-Za-z]*[-]?(\d{2,4})', clean)
            if m:
                mo_str = m.group(1).lower()
                yr_str = m.group(2)
                year = int('20' + yr_str) if len(yr_str) == 2 else int(yr_str)
                if mo_str in MO:
                    mo = MO[mo_str]
                    for j in range(3):
                        cm = mo + j
                        cy = year
                        if cm > 12: cm -= 12; cy += 1
                        col_month[col + j] = f'{cy:04d}-{cm:02d}'
                    i += 3
                else:
                    i += 1
            else:
                i += 1
        else:
            i += 1
    return col_month

# ── 4. Parsear un archivo Excel ──────────────────────────────────
def parse_file(fpath):
    fname  = os.path.basename(fpath)
    fnorm  = norm(fname)
    results = []

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        print(f"  [ERROR] No se pudo abrir {fname}: {e}")
        return []

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    sheet_key = norm(sheet_name)

    # Keywords para lookup: del nombre de archivo y del nombre de hoja
    file_keys = []
    for kw in ['country','zitizen','primavera','gica','bosketo',
               'camino','ambarte','bella','uau5','summitgrand','grand','summit',
               'lmina','bosques']:
        if kw in fnorm or kw in sheet_key:
            file_keys.append(kw)
    # Summit Grand antes de Summit para evitar falsos positivos
    if 'summitgrand' in file_keys and 'summit' in file_keys:
        file_keys = [k for k in file_keys if k != 'summit']

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Encontrar fila de encabezado
    hdr_idx = tot_col = sold_col = disp_col = None
    for i, row in enumerate(rows[:10]):
        if 'Unidades' in row:
            hdr_idx = i
            for j, v in enumerate(row):
                if v == 'Unidades': tot_col  = j + 1
                if v == 'Vendidas': sold_col = j + 1
                if v == 'Disponibles': disp_col = j + 1
            break

    if hdr_idx is None or tot_col is None:
        print(f"  [SKIP] Sin encabezado Unidades/Vendidas: {fname}")
        return []

    # Mapa columna → mes
    col_month = build_col_month(rows[hdr_idx], disp_col)  # disp_col es 1-based, pasamos como 0-based index
    # Corregir: disp_col es 1-based, necesitamos 0-based para build_col_month
    col_month = build_col_month(rows[hdr_idx], disp_col)  # disp_col - 1 + 1 = disp_col en 1-based → pasar disp_col como idx 0-based
    # Fix the call:
    col_month = {}
    _hdr = rows[hdr_idx]
    _i = disp_col  # disp_col is 1-based, so index disp_col means column disp_col+1... let's redo cleanly
    # disp_col = column number (1-based) of 'Disponibles'
    # data starts at column disp_col+1, which is index disp_col in 0-based
    col_month = build_col_month(_hdr, disp_col)  # pass disp_col as 0-based start index → that's col disp_col+1 (1-based)

    def row_val(row, col_1based):
        idx = col_1based - 1
        return row[idx] if idx < len(row) else None

    def is_tower_row(row):
        tv = row_val(row, tot_col)
        sv = row_val(row, sold_col)
        if not (isinstance(tv, (int,float)) and not isinstance(tv, bool)): return False
        if not (isinstance(sv, (int,float)) and not isinstance(sv, bool)): return False
        if tv < 5 or tv > 1000: return False        # sanity: unit counts
        if sv < 0 or sv > tv + 5: return False
        return True

    def get_tower_name(row):
        """Busca el nombre de torre en columnas a la izquierda de tot_col."""
        candidates = []
        for j in range(min(tot_col - 1, 5)):
            v = row[j]
            if isinstance(v, str) and v.strip():
                n = norm(v)
                if n not in NOT_TOWER and len(n) >= 2:
                    candidates.append(v.strip())
        return candidates[-1] if candidates else None  # tomar el más cercano a los números

    def extract_rhythm(row):
        rh = {}
        for j, v in enumerate(row):
            col = j + 1
            if col in col_month and isinstance(v, (int,float)) and not isinstance(v,bool) and v > 0:
                rh[col_month[col]] = int(v)
        return rh

    for i, row in enumerate(rows):
        if i <= hdr_idx:
            continue
        if not is_tower_row(row):
            continue

        tname = get_tower_name(row)
        if not tname:
            continue

        tot  = int(row_val(row, tot_col))
        sold = int(row_val(row, sold_col))

        # Saltar filas TOTAL
        if norm(tname) in NOT_TOWER:
            continue

        # App ID
        app_id = resolve_tower_id(file_keys, tname)
        if not app_id:
            print(f"  ? Sin mapeo: archivo='{fname}' torre='{tname}' keys={file_keys}")
            continue

        # Ritmo de ventas (datos mensuales en la fila de torre)
        rhythm = extract_rhythm(row)

        # Buscar filas C/E (construcción/entregas) y Cron. entregas en las siguientes filas
        delivery = {}
        constr_cols = []
        e_cols = []
        # Scan up to 12 rows forward, stop at next tower row
        for k in range(1, 13):
            if i + k >= len(rows): break
            nrow = rows[i + k]
            if is_tower_row(nrow):
                break  # llegamos a la siguiente torre

            # ¿Es fila de marcadores C/E? (todas las celdas no-nulas son 'C', 'E' o 'V')
            ce_vals = [v for v in nrow if v is not None]
            if ce_vals and all(v in ('C', 'E', 'V') for v in ce_vals):
                for j, v in enumerate(nrow):
                    if v == 'C':
                        constr_cols.append(j + 1)  # columna 1-based
                    elif v == 'E':
                        e_cols.append(j + 1)       # columna E = lista para entrega

            # ¿Es fila de entregas? — buscar en TODAS las celdas de texto de la fila
            row_text = ' '.join(str(v).strip() for v in nrow if v is not None and isinstance(v, str)).lower()
            if ('cron' in row_text or 'entregas' in row_text or
                'cr\u00e9dito' in row_text or 'credito' in row_text):
                for j, v in enumerate(nrow):
                    col = j + 1
                    if col in col_month and isinstance(v,(int,float)) and not isinstance(v,bool) and v > 0:
                        delivery[col_month[col]] = int(v)

        # Fechas de construcción: meses C (obra) y E (lista para entrega)
        constr = None
        if constr_cols:
            c_months = sorted(col_month[c] for c in constr_cols if c in col_month)
            e_months = sorted(col_month[c] for c in e_cols if c in col_month)
            if c_months:
                constr = {'start': c_months[0], 'end': c_months[-1]}
                if e_months:
                    constr['eEnd'] = e_months[-1]

        parts = []
        if rhythm:   parts.append(f"ventas={sum(rhythm.values())}")
        if delivery: parts.append(f"entregas={sum(delivery.values())}")
        if constr:   parts.append(f"constr={constr['start']}→{constr['end']}")
        status_str = ' | '.join(parts) if parts else 'sin datos mensuales'

        # Fecha de lanzamiento = primer mes con ritmo de ventas proyectado,
        # SOLO si el proyecto aún no ha iniciado ventas (sold == 0).
        # Si ya tiene unidades vendidas, el lanzamiento ya ocurrió: no crear
        # hitos de lanzamiento con fechas falsas basadas en remanentes de venta.
        lanzamiento = min(rhythm.keys()) if rhythm and sold == 0 else None

        print(f"  ✓ '{tname}' → {app_id}: sold={sold}/{tot} | {status_str}")
        results.append({'id': app_id, 'sold': sold, 'tot': tot, 'rhythm': rhythm,
                        'delivery': delivery, 'constr': constr, 'lanzamiento': lanzamiento})

    return results

# ── 5b. Hitos automáticos desde reglas ─────────────────────────
import hashlib as _hl, json as _rjson

RULES_FILE = os.path.join(FOLDER, 'reglas_hitos.json')
AUTO_ID_MIN, AUTO_ID_MAX = 100000, 999999   # rango reservado para hitos automáticos

def _stable_id(tid, tipo):
    """ID numérico determinístico para un (torre, tipo) — siempre el mismo entre syncs."""
    h = int(_hl.md5(f"{tid}_{tipo}".encode()).hexdigest(), 16)
    return h % (AUTO_ID_MAX - AUTO_ID_MIN + 1) + AUTO_ID_MIN

def _subtract_months(ym, n):
    """Resta n meses a una cadena 'YYYY-MM'. Devuelve 'YYYY-MM'."""
    y, m = int(ym[:4]), int(ym[5:7])
    m -= n
    while m <= 0:
        m += 12; y -= 1
    return f"{y:04d}-{m:02d}"

def _add_months(ym, n):
    """Suma n meses a 'YYYY-MM'."""
    y, m = int(ym[:4]), int(ym[5:7])
    m += n
    while m > 12:
        m -= 12; y += 1
    return f"{y:04d}-{m:02d}"

def _calc_pe_date(sold, tot, rhythm):
    """Mes proyectado en que sold alcanza el 70% de tot, usando el ritmo de ventas."""
    if not rhythm:
        return None
    tgt = tot * 0.7
    if sold >= tgt:
        return 'already'
    c = sold
    from datetime import date as _d
    cur = _d.today().strftime('%Y-%m')
    for ym in sorted(rhythm.keys()):
        if ym <= cur:
            continue
        c += rhythm[ym]
        if c >= tgt:
            return ym
    return None

def _calc_entrega_date(constr, delivery):
    """Primer mes de entrega: eEnd de constr, o primera clave de delivery, o end de constr."""
    if constr:
        if constr.get('eEnd'):
            return constr['eEnd']
        if constr.get('end'):
            return constr['end']
    if delivery:
        return sorted(delivery.keys())[0]
    return None

def _read_existing_milestones(content):
    """Lee todos los hitos actuales del HTML y los separa en manuales vs automáticos."""
    manuales, auto = [], []
    for m in re.finditer(
        r"\{id:(\d+),tid:'([^']+)',type:'([^']*)',date:'([^']*)',desc:'([^']*)'\}",
        content):
        mid = int(m.group(1))
        obj = {'id': mid, 'tid': m.group(2), 'type': m.group(3),
               'date': m.group(4), 'desc': m.group(5)}
        if AUTO_ID_MIN <= mid <= AUTO_ID_MAX:
            auto.append(obj)
        else:
            manuales.append(obj)
    return manuales, auto

def _write_milestones(content, milestones):
    """Reemplaza el bloque milestones:[...] en el HTML."""
    lines = []
    for m in milestones:
        lines.append(
            f"    {{id:{m['id']},tid:'{m['tid']}',type:'{m['type']}',date:'{m['date']}',desc:'{m['desc']}'}},")
    new_block = "  milestones:[\n" + "\n".join(lines) + "\n  ],"
    # Buscar el bloque completo con conteo de corchetes (evita DOTALL con corchetes anidados)
    start = content.find('  milestones:[')
    if start == -1:
        print("  [WARN] Bloque 'milestones' no encontrado en el HTML")
        return content
    bracket_pos = content.index('[', start)
    depth = 0
    end = None
    for i in range(bracket_pos, len(content)):
        if content[i] == '[': depth += 1
        elif content[i] == ']':
            depth -= 1
            if depth == 0:
                end = i
                break
    if end is None:
        print("  [WARN] No se encontró el cierre del bloque 'milestones'")
        return content
    # end+1 apunta al ']', avanzar a la ',' que sigue
    tail = end + 1
    while tail < len(content) and content[tail] in (' ', '\n', '\r', '\t'): tail += 1
    if tail < len(content) and content[tail] == ',': tail += 1
    return content[:start] + new_block + content[tail:]

def _parse_periodo(texto):
    """
    Parsea el texto de 'Periodo de referencia'.
    Devuelve (referencia, meses_antes):
      referencia  : 'PE' | 'entrega' | 'lanzamiento' | None
      meses_antes : int positivo = antes, 0 = mismo mes, negativo = después
    """
    if not texto or not str(texto).strip():
        return None, 0
    s = str(texto).strip().lower()

    # "X meses antes del PE / de PE / del punto de equilibrio"
    m = re.search(r'(\d+)\s+mes(?:es)?\s+antes\s+del?\s+(?:pe\b|punto)', s)
    if m: return 'PE', int(m.group(1))

    # "X meses antes de(l|la) entrega[s]"
    m = re.search(r'(\d+)\s+mes(?:es)?\s+antes\s+de(?:l|la)?\s+entregas?', s)
    if m: return 'entrega', int(m.group(1))

    # "X meses antes del lanzamiento"
    m = re.search(r'(\d+)\s+mes(?:es)?\s+antes\s+del?\s+lanzamiento', s)
    if m: return 'lanzamiento', int(m.group(1))

    # "X meses después del PE"
    m = re.search(r'(\d+)\s+mes(?:es)?\s+despu[eé]s\s+del?\s+(?:pe\b|punto)', s)
    if m: return 'PE', -int(m.group(1))

    # "X meses después de(l|la) entrega[s]"
    m = re.search(r'(\d+)\s+mes(?:es)?\s+despu[eé]s\s+de(?:l|la)?\s+entregas?', s)
    if m: return 'entrega', -int(m.group(1))

    # "al alcance del PE" / "alcance PE" / solo "PE"
    if re.search(r'alcance.{0,10}pe\b|^pe$|^punto de equilibrio$', s): return 'PE', 0

    # "lanzamiento de ventas" / "lanzamiento"
    if 'lanzamiento' in s: return 'lanzamiento', 0

    # "al momento de la entrega" / solo "entrega"
    if 'entrega' in s: return 'entrega', 0

    return None, 0


def _read_actividades():
    """Lee 'Actividades Proyectos.xlsx' — busca en EXCEL_FOLDER (padre) y luego en FOLDER."""
    for fname in ['Actividades Proyectos.xlsx', 'actividades proyectos.xlsx',
                  'actividades_proyectos.xlsx', 'Actividades_Proyectos.xlsx']:
        for base in [EXCEL_FOLDER, FOLDER]:
            fpath = os.path.join(base, fname)
            if os.path.exists(fpath):
                break
        else:
            continue
        break
    else:
        return []

    import openpyxl as _opxl
    try:
        wb = _opxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        print(f"  [ERROR] No se pudo abrir Actividades Proyectos.xlsx: {e}")
        return []

    ws = wb.active
    actividades = []
    current_scope = 'etapa'  # default si no hay encabezado de sección
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        act    = str(row[0]).strip()
        b_empty = not (len(row) > 1 and row[1])
        c_empty = not (len(row) > 2 and row[2])
        # Detectar encabezados de sección (solo col A, B y C vacías)
        if b_empty and c_empty:
            act_up = act.upper().replace(' ', '')
            if 'PORPROYECTO' in act_up:
                current_scope = 'proyecto'
                continue
            if 'PORETAPA' in act_up:
                current_scope = 'etapa'
                continue
        resp   = str(row[1]).strip() if not b_empty else ''
        periodo= str(row[2]).strip() if not c_empty else ''
        if act:
            actividades.append({'nombre': act, 'responsable': resp,
                                'periodo': periodo, 'scope': current_scope})
    return actividades


def _upsert_owners(owner_updates):
    """Escribe/actualiza responsables en Supabase (task_owners) sin tocar otros campos."""
    if not owner_updates:
        return
    import json as _jj
    from urllib.request import Request as _R2, urlopen as _U2
    url  = f'{SB_URL_E}/rest/v1/task_owners'
    hdrs = {
        'apikey': SB_KEY_E, 'Authorization': f'Bearer {SB_KEY_E}',
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates,return=minimal',
    }
    records = [{'task_id': tid, 'owner': owner} for tid, owner in owner_updates.items()]
    batch_size = 50
    total_ok = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        data  = _jj.dumps(batch).encode()
        req   = _R2(url, data=data, headers=hdrs, method='POST')
        try:
            with _U2(req, timeout=10): pass
            total_ok += len(batch)
        except Exception as e:
            print(f"  ⚠️  Supabase owners: {e}")
    if total_ok:
        print(f"  ✅ {total_ok} responsables sincronizados en Supabase")


def _js_safe(s):
    """Escapa comillas simples para uso dentro de strings JS."""
    return str(s).replace("'", "\\'").replace('\n', ' ')


def update_milestones_from_excel(content, data):
    """
    Lee 'Actividades Proyectos.xlsx' y genera un hito por actividad×torre.
    - IDs en [100000, 999999]: se regeneran en cada sync (nunca manuales).
    - Hitos manuales (id < 100000) se conservan siempre.
    - Responsables se escriben en Supabase automáticamente.
    - Si no existe el Excel, cae al JSON reglas_hitos.json como fallback.
    """
    actividades = _read_actividades()
    if not actividades:
        # Fallback al JSON anterior
        return _update_milestones_json_fallback(content, data)

    manuales, _ = _read_existing_milestones(content)
    nuevos_auto = []
    owner_updates = {}
    sin_fecha = 0

    # Agrupar torres por proyecto para actividades POR PROYECTO
    towers_list = _extract_towers()
    tid_to_proj = {t['id']: (t.get('proj') or t['id']) for t in towers_list}
    proj_tids = {}
    for tid in sorted(data.keys()):
        pid = tid_to_proj.get(tid, tid)
        if pid not in proj_tids:
            proj_tids[pid] = []
        proj_tids[pid].append(tid)

    def _proj_ref_date(pid, referencia):
        """Fecha de referencia agregada para el proyecto: mínima entre todas sus etapas."""
        dates = []
        for t in proj_tids.get(pid, []):
            d = data[t]
            if referencia == 'PE':
                dt = _calc_pe_date(d.get('sold', 0), d.get('tot', 0), d.get('rhythm', {}))
                if dt and dt != 'already': dates.append(dt)
            elif referencia == 'entrega':
                dt = _calc_entrega_date(d.get('constr'), d.get('delivery', {}))
                if dt: dates.append(dt)
            elif referencia == 'lanzamiento':
                dt = d.get('lanzamiento')
                if dt: dates.append(dt)
        return min(dates) if dates else None

    for act in actividades:
        referencia, meses_antes = _parse_periodo(act['periodo'])
        tipo  = _js_safe(act['nombre'])
        scope = act.get('scope', 'etapa')

        if scope == 'proyecto':
            # Un hito por proyecto (no por torre)
            for pid, p_tids in sorted(proj_tids.items()):
                ref_date = _proj_ref_date(pid, referencia) if referencia else None
                if ref_date is None:
                    sin_fecha += 1
                    continue
                if meses_antes > 0:
                    target_date = _subtract_months(ref_date, meses_antes)
                elif meses_antes < 0:
                    target_date = _add_months(ref_date, abs(meses_antes))
                else:
                    target_date = ref_date
                stable  = _stable_id(f'~{pid}', tipo)  # ~ separa IDs de proyecto vs torre
                rep_tid = p_tids[0]
                desc    = _js_safe(f"{act['nombre']} · {pid}")
                nuevos_auto.append({
                    'id': stable, 'tid': rep_tid,
                    'type': tipo[:35],
                    'date': target_date,
                    'desc': desc,
                })
                if act['responsable']:
                    owner_updates['m' + str(stable)] = _js_safe(act['responsable'])

        else:
            # scope == 'etapa': un hito por torre (comportamiento original)
            for tid, d in sorted(data.items()):
                if referencia == 'PE':
                    ref_date = _calc_pe_date(d.get('sold', 0), d.get('tot', 0), d.get('rhythm', {}))
                    if not ref_date or ref_date == 'already': continue
                elif referencia == 'entrega':
                    ref_date = _calc_entrega_date(d.get('constr'), d.get('delivery', {}))
                    if not ref_date: continue
                elif referencia == 'lanzamiento':
                    ref_date = d.get('lanzamiento')
                    if not ref_date: continue
                else:
                    ref_date = None

                if ref_date is None:
                    sin_fecha += 1
                    continue
                elif meses_antes > 0:
                    target_date = _subtract_months(ref_date, meses_antes)
                elif meses_antes < 0:
                    target_date = _add_months(ref_date, abs(meses_antes))
                else:
                    target_date = ref_date

                stable = _stable_id(tid, tipo)
                desc   = _js_safe(f"{act['nombre']} · {tid}")
                nuevos_auto.append({
                    'id': stable, 'tid': tid,
                    'type': tipo[:35],
                    'date': target_date,
                    'desc': desc,
                })
                if act['responsable']:
                    owner_updates['m' + str(stable)] = _js_safe(act['responsable'])

    todos = manuales + sorted(nuevos_auto, key=lambda x: (x['date'] or 'z', x['tid']))
    _upsert_owners(owner_updates)

    actos_con_periodo = len(actividades) - sum(1 for a in actividades if not _parse_periodo(a['periodo'])[0])
    print(f"\n📌 Hitos generados desde Actividades Proyectos.xlsx:")
    print(f"   Hitos con fecha:  {len(nuevos_auto)}  "
          f"({actos_con_periodo} actividades × torres con fecha calculable)")
    print(f"   Actividades sin periodo aún: {sin_fecha // max(len(data),1)} "
          f"(se agregarán cuando completes la columna C)")
    print(f"   Manuales conservados: {len(manuales)}")
    return _write_milestones(content, todos)


def _update_milestones_json_fallback(content, data):
    """Fallback: usa reglas_hitos.json si no hay Excel."""
    if not os.path.exists(RULES_FILE):
        return content
    with open(RULES_FILE, encoding='utf-8') as f:
        rules = _rjson.load(f)
    manuales, _ = _read_existing_milestones(content)
    nuevos_auto = []
    for rule in rules:
        tipo      = rule.get('tipo', '')
        referencia= rule.get('referencia', 'PE').upper()
        meses     = int(rule.get('meses_antes', 0))
        torres_ok = rule.get('torres', '*')
        tmpl      = rule.get('descripcion_template', f'{tipo} {{torre}}')
        for tid, d in sorted(data.items()):
            if torres_ok != '*' and isinstance(torres_ok, list) and tid not in torres_ok:
                continue
            if referencia == 'PE':
                ref_date = _calc_pe_date(d.get('sold', 0), d.get('tot', 0), d.get('rhythm', {}))
                if not ref_date or ref_date == 'already': continue
            elif referencia in ('ENTREGA', 'ENTREGAS'):
                ref_date = _calc_entrega_date(d.get('constr'), d.get('delivery', {}))
                if not ref_date: continue
            else:
                continue
            target_date = _subtract_months(ref_date, meses)
            stable = _stable_id(tid, tipo)
            nuevos_auto.append({'id': stable, 'tid': tid, 'type': tipo,
                                'date': target_date, 'desc': tmpl.replace('{torre}', tid)})
    todos = manuales + sorted(nuevos_auto, key=lambda x: (x['date'], x['tid']))
    print(f"\n📌 Hitos (JSON fallback): {len(nuevos_auto)} generados, {len(manuales)} manuales")
    return _write_milestones(content, todos)


def _inject_password_hash(content):
    """Lee app_password de config.json, calcula su SHA-256 e inyecta en el HTML."""
    import hashlib as _hl2
    cfg_path = os.path.join(FOLDER, 'config.json')
    if not os.path.exists(cfg_path):
        return content
    try:
        with open(cfg_path, encoding='utf-8') as f:
            cfg = _json.load(f)
        pw = cfg.get('app_password', '')
        if not pw:
            return content
        h = _hl2.sha256(pw.encode()).hexdigest()
        # Reemplaza la constante PW_HASH en el HTML
        pattern = r"const PW_HASH='[^']*'"
        if not re.search(pattern, content):
            print("  [WARN] PW_HASH no encontrado — asegúrate de tener la pantalla de login en el HTML")
            return content
        new_content = re.sub(pattern, f"const PW_HASH='{h}'", content)
        print(f"  🔐 Contraseña de acceso sincronizada")
        return new_content
    except Exception as e:
        print(f"  ⚠️  Error leyendo config para login: {e}")
        return content


# ── 5. Actualizar el HTML ───────────────────────────────────────
def update_html(all_results):
    if not os.path.exists(HTML):
        print(f"\n[ERROR] No se encontró {HTML}")
        print("Asegúrate de que crocoveen.html esté en la misma carpeta que este script.")
        return False

    with open(HTML, encoding='utf-8') as f:
        content = f.read()

    # Consolidar por tower ID (último archivo gana en caso de duplicados)
    data = {}
    for r in all_results:
        tid = r['id']
        if tid not in data:
            data[tid] = r
        else:
            # Fusionar
            if r['sold'] > 0 or r['rhythm'] or r['delivery']:
                data[tid] = r

    changed_sold = []
    changed_rhythm = []
    changed_del = []

    # 5a. Actualizar sold en towers array
    def replace_sold(m):
        tid = None
        # Buscar el id en la línea completa
        id_m = re.search(r"id:'([^']+)'", m.group(0))
        if id_m:
            tid = id_m.group(1)
        if tid and tid in data:
            new_sold = data[tid]['sold']
            changed_sold.append(f"{tid}:sold={new_sold}")
            return re.sub(r"sold:\d+", f"sold:{new_sold}", m.group(0))
        return m.group(0)

    content = re.sub(
        r"\{id:'[^']+',proj:'[^']*',name:'[^']*',tot:\d+,sold:\d+\}",
        replace_sold,
        content
    )

    def replace_js_block(txt, block_name, new_content_lines):
        """Reemplaza el bloque 'block_name:{...},' en el JS usando conteo de llaves."""
        marker = f'  {block_name}:{{'
        start = txt.find(marker)
        if start == -1:
            print(f"  [WARN] Bloque '{block_name}' no encontrado")
            return txt
        brace_pos = start + len(marker) - 1  # posición de la '{'
        depth = 0
        end = None
        for i in range(brace_pos, len(txt)):
            if txt[i] == '{':  depth += 1
            elif txt[i] == '}':
                depth -= 1
                if depth == 0:
                    end = i
                    break
        if end is None:
            print(f"  [WARN] No se encontró el cierre de '{block_name}'")
            return txt
        # Avanzar hasta después de la coma que sigue al bloque
        tail = end + 1
        while tail < len(txt) and txt[tail] in (' ', '\n', '\r', '\t'):
            tail += 1
        if tail < len(txt) and txt[tail] == ',':
            tail += 1
        new_block = f"  {block_name}:{{\n" + "\n".join(new_content_lines) + "\n  },"
        return txt[:start] + new_block + txt[tail:]

    # 5b. Reconstruir bloque rhythm:{...}
    new_rhythm_lines = []
    for tid, d in sorted(data.items()):
        if d['rhythm']:
            pairs = ','.join(f"'{k}':{v}" for k, v in sorted(d['rhythm'].items()))
            new_rhythm_lines.append(f"    '{tid}':{{{pairs}}},")
            changed_rhythm.append(tid)

    content = replace_js_block(content, 'rhythm', new_rhythm_lines)

    # 5c. Reconstruir bloque deliveries:{...}
    new_del_lines = []
    for tid, d in sorted(data.items()):
        if d['delivery']:
            pairs = ','.join(f"'{k}':{v}" for k, v in sorted(d['delivery'].items()))
            new_del_lines.append(f"    '{tid}':{{{pairs}}},")
            changed_del.append(tid)

    content = replace_js_block(content, 'deliveries', new_del_lines)

    # 5d. Reconstruir bloque constr:{...}
    # Estrategia: leer el constr actual del HTML, actualizar solo las torres
    # donde el Excel tiene marcadores C; el resto se mantiene igual.
    changed_constr = []

    # Extraer constr actual del HTML (antes de reemplazarlo)
    constr_start = content.find('  constr:{')
    brace_pos = constr_start + len('  constr:{') - 1
    depth = 0
    constr_end = None
    for i in range(brace_pos, len(content)):
        if content[i] == '{':  depth += 1
        elif content[i] == '}':
            depth -= 1
            if depth == 0:
                constr_end = i
                break

    existing_constr = {}
    if constr_end:
        block_text = content[constr_start:constr_end+1]
        # Parsear líneas como: 'C-T1A':{start:'2025-12',end:'2026-06'} o con eEnd opcional
        for m in re.finditer(r"'([^']+)':\{start:'([^']+)',end:'([^']+)'(?:,eEnd:'([^']+)')?\}", block_text):
            existing_constr[m.group(1)] = {'start': m.group(2), 'end': m.group(3)}
            if m.group(4):
                existing_constr[m.group(1)]['eEnd'] = m.group(4)

    # Mezclar: usar datos del Excel donde hay marcadores C; mantener el resto
    merged_constr = dict(existing_constr)
    for tid, d in data.items():
        if d.get('constr'):
            if existing_constr.get(tid) != d['constr']:
                changed_constr.append(
                    f"{tid}: {existing_constr.get(tid,{}).get('start','?')}→"
                    f"{existing_constr.get(tid,{}).get('end','?')}  →  "
                    f"{d['constr']['start']}→{d['constr']['end']}"
                )
            merged_constr[tid] = d['constr']

    new_constr_lines = []
    for tid in sorted(merged_constr):
        c = merged_constr[tid]
        epart = f",eEnd:'{c['eEnd']}'" if c.get('eEnd') else ''
        new_constr_lines.append(f"    '{tid}':" + "{" + f"start:'{c['start']}',end:'{c['end']}'{epart}" + "},")

    content = replace_js_block(content, 'constr', new_constr_lines)

    # 5e. Hitos automáticos desde Actividades Proyectos.xlsx (o JSON como fallback)
    content = update_milestones_from_excel(content, data)

    # 5f. Actualizar currentDate al mes actual automáticamente
    from datetime import date as _today_d
    cur_month = _today_d.today().strftime('%Y-%m')
    content_new = re.sub(r"currentDate:'[0-9]{4}-[0-9]{2}'", f"currentDate:'{cur_month}'", content)
    if content_new != content:
        print(f"  📅 currentDate actualizado a {cur_month}")
    content = content_new

    # 5g. Inyectar hash de contraseña de acceso
    content = _inject_password_hash(content)

    with open(HTML, 'w', encoding='utf-8') as f:
        f.write(content)

    print(f"\n✅ HTML actualizado: {HTML}")
    print(f"   sold actualizado:    {len(changed_sold)} torres")
    print(f"   rhythm actualizado:  {len(changed_rhythm)} torres")
    print(f"   deliveries actualizado: {len(changed_del)} torres")
    print(f"   constr actualizado:  {len(changed_constr)} torres")
    if changed_constr:
        for c in changed_constr:
            print(f"     · {c}")
    return True

# ── 6. Main ─────────────────────────────────────────────────────
def main():
    # Flag --no-email: sincroniza y hace push pero NO envía correo
    no_email = '--no-email' in sys.argv

    print(f"📂 App:    {FOLDER}")
    print(f"📂 Excel:  {EXCEL_FOLDER}")
    print(f"📄 HTML:   {HTML}")
    if no_email:
        print("📧 Modo: sin correo (--no-email)\n")
    else:
        print()

    xl_files = [
        f for f in glob.glob(os.path.join(EXCEL_FOLDER, "*.xlsx"))
        if 'Crocoveen' in os.path.basename(f) or 'crocoveen' in os.path.basename(f).lower()
    ]

    if not xl_files:
        print("[ERROR] No se encontraron archivos *Crocoveen*.xlsx")
        return

    all_results = []
    for fpath in sorted(xl_files):
        fname = os.path.basename(fpath)
        print(f"\n📊 {fname}")
        results = parse_file(fpath)
        all_results.extend(results)

    print(f"\n── Total: {len(all_results)} torres parseadas ──")
    update_html(all_results)
    push_to_github(FOLDER)
    if no_email:
        print("\n📧 Correo omitido (modo --no-email).")
    else:
        send_reminder(all_results)

# ── 7. GitHub push ──────────────────────────────────────────────
def push_to_github(folder):
    """Sube crocoveen.html a GitHub Pages automáticamente."""
    import subprocess
    print("\n🚀 Subiendo a GitHub...")
    # Verificar que git esté instalado
    try:
        subprocess.run(['git', '--version'], capture_output=True, check=True)
    except FileNotFoundError:
        print("❌  Git no está instalado.")
        print("    Descárgalo en: https://git-scm.com/download/win")
        return
    # Verificar que la carpeta sea un repositorio git
    result = subprocess.run(['git', 'rev-parse', '--git-dir'],
                            capture_output=True, cwd=folder)
    if result.returncode != 0:
        print("⚠️   Esta carpeta no está conectada a GitHub todavía.")
        print("    Sigue la Guía de Configuración Inicial (GitHub_Setup.docx).")
        return
    # Agregar y commitear
    from datetime import datetime
    fecha = datetime.now().strftime('%Y-%m-%d')
    subprocess.run(['git', 'add', 'crocoveen.html'], cwd=folder, check=True)
    # ¿Hay cambios para subir?
    sin_cambios = subprocess.run(['git', 'diff', '--cached', '--quiet'], cwd=folder)
    if sin_cambios.returncode == 0:
        print("✓  Sin cambios nuevos — GitHub ya está al día.")
        return
    try:
        subprocess.run(['git', 'commit', '-m', f'Sync {fecha}'], cwd=folder, check=True)
        subprocess.run(['git', 'push'], cwd=folder, check=True)
        print(f"✓  Publicado en GitHub ({fecha}) — el sitio se actualiza en ~1 minuto.")
    except subprocess.CalledProcessError:
        print("❌  Error al subir. Verifica tu conexión y que hayas hecho la configuración inicial.")

# ── 8. Email recordatorio ────────────────────────────────────────
import smtplib, ssl, json as _json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from urllib.request import Request as _Req, urlopen as _urlopen
from datetime import date as _date, timedelta as _td

SB_URL_E = 'https://vsxqlyhrakxqiwtdevsj.supabase.co'
SB_KEY_E = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZzeHFseWhyYWt4cWl3dGRldnNqIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzQzNjAzNjksImV4cCI6MjA4OTkzNjM2OX0.k5Mn7g8oy796qs64Sr_jLRRjZWlGEgaiawExHpxj4nI'

EMAIL_TO = [
    'jisaza@tierragrata.co',
    'andresmesab@tierragrata.co',
    'vroldan@tierragrata.co',
    'jbarbosa@tierragrata.co',
    'jsuarez@tierragrata.co',
    'jarango@tierragrata.co',
    'nparra@tierragrata.co',
    'chincapie@tierragrata.co',
    'aestrada@tierragrata.co',
    'dsierra@tierragrata.co',
]

def _sb_get(table, qs=''):
    url = f'{SB_URL_E}/rest/v1/{table}' + (f'?{qs}' if qs else '')
    req = _Req(url, headers={'apikey': SB_KEY_E, 'Authorization': f'Bearer {SB_KEY_E}'})
    try:
        with _urlopen(req, timeout=10) as r:
            return _json.loads(r.read())
    except Exception as e:
        print(f"  ⚠️  Supabase: {e}")
        return []

def _extract_towers():
    try:
        with open(HTML, encoding='utf-8') as f:
            content = f.read()
        m = re.search(r'towers:\s*\[([^\]]*)\]', content, re.DOTALL)
        if not m: return []
        towers = []
        for tm in re.finditer(r'\{[^}]+\}', m.group(1)):
            o = tm.group(0)
            t = {}
            for field in ['id','name','proj']:
                fm = re.search(rf"{field}:'([^']*)'", o)
                if fm: t[field] = fm.group(1)
            for field in ['tot','sold']:
                fm = re.search(rf"{field}:(\d+)", o)
                if fm: t[field] = int(fm.group(1))
            if 'id' in t: towers.append(t)
        return towers
    except: return []

def _extract_milestones():
    try:
        with open(HTML, encoding='utf-8') as f:
            content = f.read()
        items = re.findall(
            r"\{id:(\d+),tid:'([^']+)',type:'([^']*)',date:'([^']*)',desc:'([^']*)'\}",
            content)
        return [{'id':i,'tid':tid,'type':tp,'date':dt,'desc':ds}
                for i,tid,tp,dt,ds in items]
    except: return []

def _build_tasks():
    towers   = {t['id']: t for t in _extract_towers()}
    tasks    = []
    for m in _extract_milestones():
        t = towers.get(m['tid'], {})
        tasks.append({'id':'m'+m['id'],'tid':m['tid'],
                      'proj':t.get('proj',''),'name':m['desc'] or m['type'],
                      'type':m['type'],'date':m['date'],'tower':t.get('name',m['tid'])})
    for t in towers.values():
        if t.get('sold',0) < t.get('tot',0)*0.7:
            tasks.append({'id':'pe_'+t['id'],'tid':t['id'],'proj':t.get('proj',''),
                          'name':'70% en ventas','type':'PE','date':None,'tower':t.get('name',t['id'])})
    return tasks

def _mes(ym):
    if not ym: return ''
    MN=['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    try: y,m=ym.split('-'); return f"{MN[int(m)-1]}-{y[2:]}"
    except: return ym

def send_reminder(all_results):
    cfg_path = os.path.join(FOLDER, 'config.json')
    if not os.path.exists(cfg_path):
        print("\n📧 Email: crea config.json con tus credenciales para activar recordatorios.")
        return
    with open(cfg_path, encoding='utf-8') as f:
        cfg = _json.load(f)
    email_from = cfg.get('email_from','')
    email_pass = cfg.get('email_password','')
    if not email_from or not email_pass or email_pass == 'TU_CONTRASEÑA':
        print("\n📧 Email: completa email_from y email_password en config.json")
        return

    print("\n📧 Preparando correo de recordatorio...")
    comp    = _sb_get('task_completions','done=eq.true&select=task_id,completed_at')
    own_raw = _sb_get('task_owners','select=task_id,owner')
    done_ids= {r['task_id'] for r in comp}
    done_at = {r['task_id']: (r.get('completed_at') or '')[:10] for r in comp}
    owners  = {r['task_id']: r['owner'] for r in own_raw}

    tasks   = _build_tasks()
    today   = _date.today()
    w_ago   = (today - _td(days=7)).isoformat()[:7]
    w_ahead = (today + _td(days=7)).isoformat()[:7]
    today_s = today.isoformat()[:7]

    pending   = [t for t in tasks if t['id'] not in done_ids]
    upcoming  = [t for t in pending if t.get('date') and today_s <= t['date'] <= w_ahead]
    comp_week = [t for t in tasks if t['id'] in done_ids and done_at.get(t['id'],'') >= w_ago]

    def _owner_badge(tid):
        o = owners.get(tid,'')
        if o: return f'<span style="background:#e6eef2;color:#0D3D52;padding:2px 8px;border-radius:10px;font-size:11px">{o}</span>'
        return '<span style="color:#ccc;font-size:11px">—</span>'

    def _row3(t):
        return (f'<tr><td style="padding:7px 12px;border-bottom:1px solid #f0f0f0;font-size:13px">'
                f'{t["name"]} <span style="color:#aaa;font-size:11px">{t["tower"]}</span></td>'
                f'<td style="padding:7px 12px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#888">{_mes(t.get("date"))}</td>'
                f'<td style="padding:7px 12px;border-bottom:1px solid #f0f0f0">{_owner_badge(t["id"])}</td></tr>')

    def _row2(t):
        return (f'<tr><td style="padding:7px 12px;border-bottom:1px solid #f0f0f0;font-size:13px">'
                f'{t["name"]} <span style="color:#aaa;font-size:11px">{t["tower"]}</span></td>'
                f'<td style="padding:7px 12px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#888">{_mes(t.get("date"))}</td></tr>')

    by_proj = {}
    for t in pending:
        by_proj.setdefault(t['proj'] or 'Sin proyecto', []).append(t)
    pend_rows = ''
    for proj, ts in sorted(by_proj.items()):
        pend_rows += (f'<tr><td colspan="3" style="padding:8px 12px 3px;background:#f9f9f9;'
                      f'font-weight:700;color:#007060;font-size:11px;text-transform:uppercase">{proj}</td></tr>')
        for t in ts[:15]: pend_rows += _row3(t)

    up_rows   = ''.join(_row3(t) for t in upcoming) or '<tr><td colspan="3" style="padding:12px;color:#aaa;text-align:center">Ninguna tarea vence esta semana 🎉</td></tr>'
    comp_rows = ''.join(_row2(t) for t in comp_week[:15]) or '<tr><td colspan="2" style="padding:12px;color:#aaa;text-align:center">Ninguna completada esta semana</td></tr>'

    MES_ES = {'January':'enero','February':'febrero','March':'marzo','April':'abril',
              'May':'mayo','June':'junio','July':'julio','August':'agosto',
              'September':'septiembre','October':'octubre','November':'noviembre','December':'diciembre'}
    fecha_str = today.strftime('%d de %B de %Y')
    for en,es in MES_ES.items(): fecha_str = fecha_str.replace(en,es)

    TH = 'style="padding:7px 12px;text-align:left;font-size:11px;color:#888;font-weight:600;border-bottom:2px solid #eee"'
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f5f5f5;font-family:Arial,sans-serif">
<div style="max-width:640px;margin:32px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.08)">
  <div style="background:#0D3D52;padding:28px 32px">
    <div style="font-size:22px;font-weight:700;color:#fff">🌿 Crocoveen</div>
    <div style="color:#a8cdd8;font-size:13px;margin-top:4px">Resumen semanal · {fecha_str}</div>
  </div>
  <div style="display:flex;padding:18px 32px;gap:12px;background:#eaf2f5;border-bottom:1px solid #c8dde5">
    <div style="flex:1;text-align:center"><div style="font-size:26px;font-weight:700;color:#0D3D52">{len(pending)}</div><div style="font-size:11px;color:#666">Pendientes</div></div>
    <div style="flex:1;text-align:center"><div style="font-size:26px;font-weight:700;color:#e65100">{len(upcoming)}</div><div style="font-size:11px;color:#666">Vencen esta semana</div></div>
    <div style="flex:1;text-align:center"><div style="font-size:26px;font-weight:700;color:#007060">{len(comp_week)}</div><div style="font-size:11px;color:#666">Completadas (7 días)</div></div>
  </div>
  <div style="padding:22px 32px 0">
    <div style="font-size:14px;font-weight:700;color:#333;margin-bottom:10px">⚡ Vencen esta semana</div>
    <table style="width:100%;border-collapse:collapse">
      <tr style="background:#fff8e1"><th {TH}>TAREA</th><th {TH}>FECHA</th><th {TH}>RESPONSABLE</th></tr>
      {up_rows}
    </table>
  </div>
  <div style="padding:22px 32px 0">
    <div style="font-size:14px;font-weight:700;color:#333;margin-bottom:10px">📋 Pendientes por proyecto</div>
    <table style="width:100%;border-collapse:collapse">
      <tr style="background:#f5f5f5"><th {TH}>TAREA</th><th {TH}>FECHA</th><th {TH}>RESPONSABLE</th></tr>
      {pend_rows or '<tr><td colspan="3" style="padding:12px;color:#aaa;text-align:center">No hay tareas pendientes 🎉</td></tr>'}
    </table>
  </div>
  <div style="padding:22px 32px">
    <div style="font-size:14px;font-weight:700;color:#333;margin-bottom:10px">✅ Completadas esta semana</div>
    <table style="width:100%;border-collapse:collapse">
      <tr style="background:#eaf2f5"><th {TH}>TAREA</th><th {TH}>FECHA</th></tr>
      {comp_rows}
    </table>
  </div>
  <div style="padding:0 32px 28px;text-align:center">
    <a href="https://julianaisaza.github.io/Crocoveen" style="display:inline-block;background:#0D3D52;color:#fff;text-decoration:none;padding:11px 26px;border-radius:8px;font-weight:600;font-size:13px">Ver Crocoveen →</a>
  </div>
  <div style="background:#f5f5f5;padding:14px 32px;text-align:center">
    <div style="font-size:11px;color:#aaa">Correo automático semanal · Tierra Grata & Co.</div>
  </div>
</div></body></html>"""

    msg = MIMEMultipart('alternative')
    msg['Subject'] = f'🌿 Crocoveen — Resumen semanal {today.strftime("%d/%m/%Y")}'
    msg['From']    = email_from
    msg['To']      = ', '.join(EMAIL_TO)
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    try:
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=ctx) as s:
            s.login(email_from, email_pass)
            s.sendmail(email_from, EMAIL_TO, msg.as_string())
        print(f"✓  Correo enviado a {len(EMAIL_TO)} personas.")
    except Exception as e:
        print(f"❌  Error enviando correo: {e}")
        print("   Verifica email_from y email_password en config.json")

# ── 9. Main ─────────────────────────────────────────────────────
if __name__ == '__main__':
    main()
